import logging
from multiprocessing.pool import ThreadPool
import os
import json
import subprocess
from dataclasses import asdict
import base64

import pypdfium2 as pdfium
from openpyxl import load_workbook
import PIL

from llmonkey.llmonkey import LLMonkey
from llmonkey.models import ModelProvider
from llmonkey.providers.openai_like import IonosProvider, MistralProvider

from ingest.schemas import UploadedOrExtractedFile, AiMetadataResult
from ingest.logic.common import UPLOADED_FILES_FOLDER


def import_office_document(files: list[UploadedOrExtractedFile], parameters, on_progress=None) -> tuple[list[dict], list[dict]]:
    from pdferret import PDFerret
    from pdferret.datamodels import ChunkType, PDFChunk, PDFDoc
    import nltk
    nltk.download('punkt')

    extractor = PDFerret(meta_extractor="grobid", text_extractor="unstructured", general_extractor="tika", llm_summary=False, llm_table_description=False)

    if on_progress:
        on_progress(0.1)
    parsed, failed = extractor.extract_batch([f'{UPLOADED_FILES_FOLDER}/{uploaded_file.local_path}' for uploaded_file in files])
    failed_files = [{"filename": pdferror.file, "reason": pdferror.exc} for pdferror in failed]

    if on_progress:
        on_progress(0.5)

    items = []

    def postprocess_item(parsed_file, uploaded_file):
        file_metainfo = parsed_file.metainfo
        # MetaInfo({'doi': '', 'title': ['Protokoll', 'Druck'], 'abstract': '',
        # 'authors': ['Georg T...', 'Klaus H...'], 'pub_date': ['2023-04-05T12:56:00Z',
        # '2023-04-11T07:13:20Z', '2013-08-07T11:41:02'], 'language': '',
        # 'file_features': FileFeatures(filename='...',
        # file='...', is_scanned=None), 'npages': None, 'thumbnail': b'\x89PNG...

        if file_metainfo.file_features.filename.endswith((".docx")):
            import pypandoc
            markdown = pypandoc.convert_file(file_metainfo.file_features.file,
                                 'markdown-escaped_line_breaks',
                                 format='docx', extra_args=['--columns=130'])
            chunks = []
            line_chunk = []
            for line in markdown.split("\n"):
                line_chunk.append(line)
                if len(line_chunk) == 12:
                    chunk = PDFChunk(text="\n".join(line_chunk), chunk_type=ChunkType.TEXT)
                    chunks.append(chunk)
                    line_chunk = []
            if line_chunk:
                chunk = PDFChunk(text="\n".join(line_chunk), chunk_type=ChunkType.TEXT)
                chunks.append(chunk)
            parsed_file.chunks = chunks
        elif file_metainfo.file_features.filename.endswith((".xls", ".xlsx")):
            for chunk in parsed_file.chunks:
                if chunk.chunk_type == ChunkType.TABLE:
                    chunk.text = chunk.non_embeddable_content[:800]

            sheetnames = get_sheet_names(file_metainfo.file_features.file)
            text = "This Excel spreadsheet contains the following sheets: " + ", ".join(sheetnames)
            text_de = "Diese Excel-Tabelle enthält die folgenden Arbeitsblätter: " + ", ".join(sheetnames)
            logging.warning(f"Prepending sheet names to text: {text_de}")
            chunk = PDFChunk(text=text_de, chunk_type=ChunkType.TEXT)
            parsed_file.chunks = [chunk] + parsed_file.chunks
        elif file_metainfo.file_features.filename.endswith((".pdf", )):
            thumbnail = file_metainfo.thumbnail
            if thumbnail:
                url_encoded = "data:image/png;base64," + base64.b64encode(thumbnail).decode("utf-8")
                text_de = """Du erhälst die erste Seite eines PDF-Dokuments. Fasse das Dokument in bis zu drei Sätzen zusammen."""
                prompt = [
                    {
                        "type": "text",
                        "text": text_de
                    },
                    {
                        "type": "image_url",
                        "image_url": url_encoded
                    }
                ]
                llmonkey_instance = LLMonkey()
                response = llmonkey_instance.generate_prompt_response(
                    provider=ModelProvider.mistral,
                    model_name=MistralProvider.Models.PIXTRAL.identifier,
                    user_prompt=prompt,
                    max_tokens=2000,
                )
                response_text = response.conversation[-1].content

                assert isinstance(response_text, str)
                summary_chunk = PDFChunk(text="AI Summary: " + response_text, chunk_type=ChunkType.TEXT)
                #text_chunk = PDFChunk(text=info['text'], chunk_type=ChunkType.TEXT)
                parsed_file.chunks = [summary_chunk] + parsed_file.chunks

        if not file_metainfo.title and not len(parsed_file.chunks):
            failed_files.append({"filename": uploaded_file.local_path, "reason": "no title or text found, skipping"})
            return None

        full_text_chunks = [asdict(chunk) for chunk in parsed_file.chunks]
        for chunk in full_text_chunks:
            if len(chunk['non_embeddable_content']) > 5000:
                chunk['non_embeddable_content'] = chunk['non_embeddable_content'][:5000] + "..."
            if len(chunk['text']) > 10000:
                chunk['text'] = chunk['text'][:10000] + "..."
            chunk.pop('chunk_type')
        full_text = " ".join([chunk['text'] for chunk in full_text_chunks])

        ai_metadata = get_ai_summary(uploaded_file.original_filename,
                                     uploaded_file.metadata.folder if uploaded_file.metadata else None,
                                     full_text)

        folder = uploaded_file.metadata.folder if uploaded_file.metadata else None
        item = {
            "title": (ai_metadata and ai_metadata.title) or uploaded_file.original_filename,
            "type_description": (ai_metadata and ai_metadata.document_type) or None,
            "abstract": (ai_metadata and ai_metadata.summary) or file_metainfo.abstract,
            "ai_tags": (ai_metadata and ai_metadata.tags) or [],
            "content_date": (ai_metadata and ai_metadata.date) or None,
            "content_time": (ai_metadata and ai_metadata.time) or None,
            "language": (ai_metadata and ai_metadata.document_language) or file_metainfo.language or parameters.get("default_language", "de"),
            "thumbnail_path": store_thumbnail(file_metainfo.thumbnail, uploaded_file.local_path) if file_metainfo.thumbnail else None,  # relative to UPLOADED_FILES_FOLDER

            "full_text": full_text,
            "full_text_chunks": full_text_chunks,

            "file_created_at": uploaded_file.metadata.created_at if uploaded_file.metadata else None,
            "file_updated_at": uploaded_file.metadata.updated_at if uploaded_file.metadata else None,
            "file_name": uploaded_file.original_filename,
            "file_type": uploaded_file.local_path.split(".")[-1],
            "folder": folder,
            "full_path": os.path.join(folder or "", uploaded_file.original_filename),
            "uploaded_file_path": uploaded_file.local_path,  # relative to UPLOADED_FILES_FOLDER
        }
        return item

    with ThreadPool(processes=10) as pool:
        items = pool.map(lambda args: postprocess_item(args[0], args[1]), zip(parsed, files))

    items = [item for item in items if item]

    return items, failed_files


def get_sheet_names(file_path):
    try:
        # Load the workbook without loading all data
        workbook = load_workbook(filename=file_path, read_only=True)
        # Get the list of sheet names
        sheet_names = workbook.sheetnames
        return sheet_names
    except Exception as e:
        logging.warning(f"Failed to get sheet names from {file_path}: {e}")
        return []


def get_ai_summary(title: str, folder: str | None, full_text: str) -> AiMetadataResult | None:
    prompt = """
You are an expert in extracting titles and short summaries from documents.
The available information about the document start with <document> and end with </document>.

<document>
Folder: {{ folder }}
File Name: {{ title }}
Content: {{ content }}
</document>

Try to extract the title (keep it close to what is mentioned in the document, and its not necessarily the file name), the abstract document type and a one-sentence summary of the content as a JSON object.
The abstract document type should be something like: "paper about a new algorithm", "report about a waste plant", "presentation about a new product".
The summary should be a single sentence that captures the essence of the document.
Leave any field empty if you can't find the information.

Schema:
{
    "document_language": "Two-letter code of the document language",
    "title": "Title of the document",
    "document_type": "Abstract document type",
    "summary": "Two-sentence summary of the content",
    "tags": ["list", "of", "tags", "related", "to", "the", "document"],
    "date": "Date of the document (if mentioned) in YYYY-MM-DD format",
    "time": "Time of the document (if mentioned) in HH:MM:SS format"
}

Reply only with the json object.
    """

    prompt_de = """
Du bist ein Experte im Extrahieren von Titeln und kurzen Zusammenfassungen aus Dokumenten.
Die verfügbaren Informationen über das Dokument beginnen mit <document> und enden mit </document>.

<document>
Ordner: {{ folder }}
Dateiname: {{ title }}
Inhalt:
{{ content }}
</document>

Extrahiere Metadaten nach folgendem JSON Schema:

{
    "document_language": "Zwei-Buchstaben-Code der Dokumentsprache, z.B. 'de' oder 'en'",
    "title": "Titel des Dokuments",
    "document_type": "Abstrakter Dokumenttyp",
    "summary": "Sehr kurze Zusammenfassung des Inhalts",
    "tags": ["Liste", "von", "Tags", "die", "mit", "dem", "Dokument", "verwandt", "sind"],
    "date": "Datum des Dokuments (falls erwähnt) im Format JJJJ-MM-TT",
    "time": "Uhrzeit des Dokuments (falls erwähnt) im Format HH:MM:SS"
}

Erläuterungen zu den Feldern:

document_language: Der Zwei-Buchstaben-Code der Dokumentsprache, z.B. 'de' oder 'en'.

title: Der Titel des Dokuments. Wenn ein Titel im Dokument genannt wird, soll er exakt so wiedergegeben werden.
Wenn kein Titel genannt wird, soll ein möglichst treffender, kurzer Titel gewählt werden.
Der Dateiname kann als Inspiration dienen, muss aber nicht übernommen werden.

document_type: Der abstrakte Dokumenttyp. Beispiele sind: "Aufsatz über einen neuen Algorithmus", "Bericht über eine Müllverbrennungsanlage", "Präsentation über ein neues Produkt".

summary: Die Zusammenfassung sollten zwei Sätze sein, die das Wesentliche des Dokuments erfassen.
Fasse die Sätz kurz, ohne Füllwörter. Konzentriere dich auf Informationen, die wichtig sein können um das Dokument zu finden, z.B. Wer, Was, Wann, Wo, Warum.

tags: Bis zu 10 Stichwörter, die das Dokument beschreiben. Nutze sowohl allgemeine Stichworte als auch spezifische Begriffe,
die im Dokument vorkommen, zum Beispiel Orte, Personen oder Technologien. Füge alle direkt beteiligten Personen als Tags hinzu.

date: Das Datum des Dokuments, falls es im Text oder Dateinamen erwähnt wird. Gib das Datum im Format JJJJ-MM-TT an. Falls nur ein Jahr oder Monat genannt wird, nutze den 01. des entsprechenden Monats.

time: Die Uhrzeit des Dokuments, falls sie im Text erwähnt wird (zum Beispiel der Startzeitpunkt eines Meetings). Gib die Uhrzeit im Format HH:MM:SS an.

Lasse ein Feld leer, wenn du die Information nicht finden kannst.
Texte und Stichwörter sollen in deutscher Sprache sein, egal in welcher Sprache das Dokument verfasst ist.
Nutze auch den Ordnername für Hinweise aus die Bedeutung eines Dokuments.

Antworte nur mit dem JSON-Objekt. Antworte mit einem vollständigen JSON-Objekt, auch wenn du nicht alle Felder ausfüllen kannst.
    """

    prompt = (prompt_de.replace("{{ title }}", title)
              .replace("{{ folder }}", folder or "")
              .replace("{{ content }}", full_text[:1000]))
    llmonkey_instance = LLMonkey()
    response = llmonkey_instance.generate_prompt_response(
        provider=ModelProvider.ionos,
        model_name=IonosProvider.Models.META_LLAMA_3_1_70B.identifier,
        system_prompt=prompt,
        max_tokens=1000,
    )
    response_text = response.conversation[-1].content
    assert response_text
    try:
        info = json.loads(response_text)
    except (KeyError, ValueError):
        logging.warning(f"Failed to parse AI response: {response_text}")
        return None
    result = AiMetadataResult(**info)
    return result


def store_thumbnail(png_data: bytes, sub_path):
    with open(f'{UPLOADED_FILES_FOLDER}/{sub_path}.thumbnail.png', 'wb') as f:
        f.write(png_data)
    return f'{sub_path}.thumbnail.png'


def get_thumbnail_office(sub_path) -> str | None:
    suffix = sub_path.split('.')[-1]
    file_path = f'{UPLOADED_FILES_FOLDER}/{sub_path}'
    folder = os.path.dirname(file_path)

    try:
        subprocess.run([
            'libreoffice', '--headless',
            '--convert-to','pdf:writer_pdf_Export:{"PageRange":{"type":"string","value":"1"}}',
            '--outdir', folder, file_path
            ], check=True)
        pdf_path = file_path.replace(suffix, 'pdf')
    except Exception as e:
        logging.warning(f"Failed to convert office document to pdf: {e}")
        return None

    try:
        pdf = pdfium.PdfDocument(pdf_path)
        first_page = pdf[0]
        image = first_page.render(scale=1).to_pil()
        thumbnail_path = f'{sub_path}.thumbnail.jpg'
        image.save(f'{UPLOADED_FILES_FOLDER}/{thumbnail_path}', 'JPEG', quality=60)
    except Exception as e:
        logging.warning(f"Failed to create thumbnail for {sub_path}: {e}")
        thumbnail_path = None
    return thumbnail_path
